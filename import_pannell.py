
from datetime import datetime, time

# pip install openpyxl
from openpyxl import load_workbook

import db

FILENAME = "Beatles song database 2024-05-27.xlsx"
COMMENT_PREFIX = "David Pannell:\n"

NOTE_TO_VARIANT = {
    "original single": "single",
    "album version": "album",
    "single version": "single",
    "charity version": "single", # ?
    "Let It Be version": "album",
    "Let It Be version": "album",
}

def split_out_variant(title):
    variant = "album"

    i = title.find(" (")
    if i >= 0 and False:
        potential_title = title[:i]
        potential_variant = NOTE_TO_VARIANT.get(title[i + 2:-1])
        if potential_variant is not None:
            title = potential_title
            variant = potential_variant

    return title, variant

def main():
    songs = db.load()

    wb = load_workbook(FILENAME, data_only=True)
    sheet = wb["Tracks"]
    header = []
    for row_number, row in enumerate(sheet.iter_rows()):
        pannell = {}
        song = None
        for col_number, cell in enumerate(row):
            if row_number == 0:
                header.append(cell.value)
                key = None
            else:
                key = header[col_number]
            if key == "Song_title":
                title = cell.value
                if title is not None:
                    title, variant = split_out_variant(title)
                    song = db.get_song_by_title(songs, title)
                    if song is None:
                        print(f"Didn't find song \"{title}\"")
                    else:
                        if "pannell" not in song:
                            song["pannell"] = {}
                        song["pannell"][variant] = pannell
            if song is not None:
                value = cell.value
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d')
                elif isinstance(value, time):
                    value = value.hour * 3600 + value.minute * 60 + value.second
                pannell[key] = value
                if key is not None and cell.comment:
                    if "comments" not in pannell:
                        pannell["comments"] = {}
                    comment = cell.comment.text
                    if comment.startswith(COMMENT_PREFIX):
                        comment = comment[len(COMMENT_PREFIX):]
                    pannell["comments"][key] = comment

    db.save(songs)

main()

